from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
import openpyxl
from datetime import datetime

def get_assignment(login_data, link):
    driver = webdriver.Chrome('chromedriver-win64/chromedriver-win64/chromedriver.exe')

    courses_info = {}

    try:
        driver.get("https://nlms.gwnu.ac.kr/login/index.php")
        
        driver.find_element(By.NAME, "username").send_keys(login_data['username'])
        driver.find_element(By.NAME, "password").send_keys(login_data['password'])
        driver.find_element(By.NAME, "loginbutton").click()

        driver.get(link)

        time.sleep(5)

        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')

        a_tag = [tag for tag in soup.select('a') if '실습' in tag.text]
        links_with_dates = []
        for tag in a_tag:
            if tag.has_attr('href'):
                parent_row = tag.find_parent('tr')
                if parent_row:
                    date_cells = parent_row.find_all('td')
                    if len(date_cells) > 1:
                        end_date = date_cells[1].text.strip()
                        links_with_dates.append((tag['href'], end_date))
        
        for info in links_with_dates:
            url = info[0]
            driver.get(url + "&action=grading")

            time.sleep(3)

            page_source = driver.page_source
            parse_assignment = BeautifulSoup(page_source, 'html.parser')
            pagination = parse_assignment.find_all('ul', class_='pagination')
            
            # 목록 페이지 개수
            page_count = 0
            if pagination:
                li_tags = pagination[0].find_all('li') # 첫번째거 사용
                for li in li_tags:
                    text = li.find('a') or li.find('span')
                    if text and text.text.isdigit():
                        page_count += 1

            # 제출 여부 정부 추출
            student_data = []
            for i in range(page_count):
                driver.get(url + f"&action=grading&page={i}")

                time.sleep(3)

                page_source = driver.page_source
                soup = BeautifulSoup(page_source, 'html.parser')
                table = soup.select("div.box.boxaligncenter.gradingtable > div.no-overflow > table > tbody")
                if table:
                    rows = table[0].select('tr')
                    for row in rows:
                        cells = row.find_all('td')
                        if len(cells) > 3:
                            student_number = cells[3].text.strip()
                            status = cells[4].text.strip()

                            if status == "제출 완료":
                                status = 'O'
                            elif "제출 완료" in status:
                                if "일" in status and "일시" in status:
                                    status = 'X'
                                else:
                                    status = 'O'
                            elif status.startswith("미제출"):
                                status = '/'
                            
                            student_data.append((student_number, status))
                
            
            courses_info[info[1]] = student_data

    finally:
        driver.quit()
    return courses_info

if __name__ == "__main__":
    id = "khjg"
    pw = "kjh!"
    link = "https://nlms.gwnu.ac.kr/mod/assign/index.php?id=13864"
    login_data = {
        'username': id,
        'password': pw
    }
    
    assignments = get_assignment(login_data, link)
    # 날짜 형식으로 키 변환
    converted_assignments = {key[:8]: value for key, value in assignments.items()}
    
    wb = openpyxl.load_workbook('응용프로그래밍 2반 출석.xlsx')
    sheet = wb.active

    date_row = 2  # 날짜가 있는 행

    # 과제 날짜와 해당 열 번호 매핑
    date_column_mapping = {}
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=date_row, column=col).value
        if isinstance(cell_value, datetime):
            date_str = cell_value.strftime('%Y%m%d')
            if date_str in converted_assignments:
                date_column_mapping[date_str] = col

    student_id_column = 1  # 학번이 있는 열 번호

    # 데이터 업데이트
    for row in range(3, sheet.max_row + 1):
        student_id = sheet.cell(row=row, column=student_id_column).value
        for date_str, student_infos in converted_assignments.items():
            if date_str in date_column_mapping:
                col = date_column_mapping[date_str]
                for student_number, status in student_infos:
                    if str(student_id) == student_number:
                        sheet.cell(row=row, column=col).value = status
                        break

    # 파일 저장
    updated_file_path = '응용프로그래밍 2반 출석_update.xlsx'
    wb.save(updated_file_path)
    wb.close()